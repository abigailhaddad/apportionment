#!/usr/bin/env python3
"""
Parse SF 133 TAFS detail data for all agencies.
Based on parse_education_sf133.py but handles multiple agencies.
"""

import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
import re

def parse_sf133_tafs_detail(file_path, agency_name=None):
    """Parse the TAFS detail sheet from an SF 133 file."""
    print(f"Processing: {file_path}")
    
    # Use openpyxl to handle merged cells properly
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Find the TAFS detail sheet
    tafs_sheet = None
    for sheet_name in wb.sheetnames:
        if 'TAFS' in sheet_name and 'detail' in sheet_name.lower():
            tafs_sheet = sheet_name
            break
    
    if not tafs_sheet:
        print(f"  No TAFS detail sheet found in {file_path}")
        wb.close()
        return None
    
    ws = wb[tafs_sheet]
    
    # Convert to dataframe, unmerging cells
    data = []
    for row in ws.iter_rows():
        data.append([cell.value for cell in row])
    
    df = pd.DataFrame(data)
    
    # Find agency name if not provided
    if not agency_name:
        for row in df.iloc[:10].values:
            for cell in row:
                if cell and isinstance(cell, str):
                    # More flexible agency detection
                    if ('Department' in str(cell) or 'Agency' in str(cell) or 
                        'Administration' in str(cell) or 'Office' in str(cell) or
                        'Corps' in str(cell) or 'Branch' in str(cell) or
                        # Specifically for Other Independent Agencies
                        str(cell).strip() == 'Other Independent Agencies'):
                        agency_name = str(cell).strip()
                        break
            if agency_name:
                break
    
    if not agency_name:
        print(f"  Could not identify agency in {file_path}")
        agency_name = "Unknown Agency"
    
    print(f"  Agency: {agency_name}")
    
    # Find the header row by looking for "Line No" or "Lineno"
    header_row = None
    for idx, row in df.iterrows():
        if any(('Line No' in str(cell) or 'Lineno' in str(cell)) if cell else False for cell in row.values):
            header_row = idx
            break
    
    if header_row is None:
        print(f"  Could not find header row in {file_path}")
        wb.close()
        return None
    
    # Set column names from header row
    df.columns = df.iloc[header_row].fillna('')
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    
    # Clean up column names
    df.columns = [str(col).strip() if col else f'Col_{i}' for i, col in enumerate(df.columns)]
    
    # Forward fill the bureau, account info, and TAFS columns
    fill_columns = ['Col_0', 'Col_1', 'Col_4']  # Bureau, Account, TAFS
    for col in fill_columns:
        if col in df.columns:
            # Only fill if the value is not a line number
            mask = df[col].astype(str).str.match(r'^\d{4}(\.\d)?$', na=False)
            df.loc[~mask, col] = df.loc[~mask, col].ffill()
    
    # Clean up the data
    df = df.replace('', np.nan)
    
    # Add agency column
    df['Agency'] = agency_name
    
    # Add source file
    df['Source_File'] = Path(file_path).name
    
    wb.close()
    
    return df

def main():
    """Process all SF 133 files and create a consolidated dataset."""
    sf133_dir = Path('raw_data/sf133')
    
    # Get all Excel files
    excel_files = list(sf133_dir.glob('*.xlsx')) + list(sf133_dir.glob('*.xls'))
    print(f"Found {len(excel_files)} Excel files to process\n")
    
    # Process each file
    all_data = []
    successful_files = 0
    agencies_processed = set()
    
    for file_path in excel_files:
        try:
            df = parse_sf133_tafs_detail(file_path)
            if df is not None and len(df) > 0:
                all_data.append(df)
                successful_files += 1
                agencies_processed.add(df['Agency'].iloc[0])
            else:
                print(f"  Skipping - no data extracted")
        except Exception as e:
            print(f"  Error processing {file_path}: {str(e)}")
        print()
    
    # Combine all data
    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        
        # Save the combined structured data
        output_path = Path('data/all_agencies_sf133_structured.csv')
        combined_df.to_csv(output_path, index=False)
        
        print(f"\n=== SUMMARY ===")
        print(f"Successfully processed: {successful_files} files")
        print(f"Total agencies: {len(agencies_processed)}")
        print(f"Agencies processed:")
        for agency in sorted(agencies_processed):
            print(f"  - {agency}")
        print(f"\nTotal rows: {len(combined_df):,}")
        print(f"Output saved to: {output_path}")
        
        # Also create a summary by agency
        agency_summary = combined_df.groupby('Agency').agg({
            'Source_File': 'first',
            'Line No': 'count'
        }).rename(columns={'Line No': 'Row_Count'})
        
        summary_path = Path('data/all_agencies_summary.csv')
        agency_summary.to_csv(summary_path)
        print(f"Agency summary saved to: {summary_path}")
        
    else:
        print("No data was successfully extracted from any files.")

if __name__ == "__main__":
    main()